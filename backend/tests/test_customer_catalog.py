import asyncio
import inspect
from io import BytesIO

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.api.customer_catalog import (
    MAX_CATALOG_PAGE_SIZE,
    ORDER_SCOPE_FIELD,
    ORDER_CHAT_TEXT_SEARCH_FIELDS,
    PART_NAME_FIELD,
    PART_SCOPE_FIELD,
    PART_STOCK_FIELD,
    PRODUCT_SORT_FIELDS,
    _english_status,
    _english_text,
    _part_inventory_workbook,
    _product_inventory_workbook,
    _customer_bom_lines,
    _order_catalog_query,
    _order_month_range,
    _part,
    _order,
    _order_details,
    _product,
    _product_detail,
    _product_image,
    _product_images,
    _related_products,
    _scoped_query,
    find_customer_orders_for_chat,
    export_customer_parts,
    export_customer_products,
    list_customer_parts,
    list_customer_orders,
    summarize_customer_orders,
)
from app.services.customer_chat_auth import CustomerSession


def test_catalog_supports_customer_page_sizes_and_all_product_table_sorts() -> None:
    assert MAX_CATALOG_PAGE_SIZE == 100
    assert set(PRODUCT_SORT_FIELDS) == {
        "productSku",
        "productName",
        "modelName",
        "scale",
        "category",
        "stock",
        "bomCount",
    }


def test_catalog_defaults_to_ten_rows_per_filemaker_request() -> None:
    from app.api.customer_catalog import list_customer_parts, list_customer_products

    product_default = inspect.signature(list_customer_products).parameters["page_size"].default
    part_default = inspect.signature(list_customer_parts).parameters["page_size"].default

    assert product_default.default == 10
    assert part_default.default == 10


def test_product_catalog_query_forces_scope_on_every_search_branch() -> None:
    query = _scoped_query(
        "MYK01",
        search_fields=("product_sku", "product_name"),
        scope_field="id_client",
        scope_value="CU638",
    )

    assert query == [
        {"product_sku": "*MYK01*", "id_client": "==CU638"},
        {"product_name": "*MYK01*", "id_client": "==CU638"},
    ]


def test_product_inventory_workbook_contains_only_sku_and_inventory() -> None:
    content = _product_inventory_workbook([
        {"fieldData": {"product_sku": "MYK-01", "stock": "12"}},
        {"fieldData": {"product_sku": "=NOT_A_FORMULA", "stock": "3.5"}},
    ])

    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Products"]
    assert list(sheet.values) == [
        ("SKU", "Inventory"),
        ("MYK-01", 12),
        ("=NOT_A_FORMULA", 3.5),
    ]
    assert sheet["A3"].data_type == "s"
    assert sheet.freeze_panes == "A2"


def test_part_inventory_workbook_contains_customer_visible_inventory_fields() -> None:
    content = _part_inventory_workbook([
        {
            "fieldData": {
                "part_number": "AL050013-00",
                "part_name_en": "Pipe Holder",
                "status": "Active",
                "stock_on_hand_qty": "8",
                "safety_stock_qty": "2",
                "Turnover Time": "12 Days",
                "Date Created": "12/05/2024",
            },
        },
        {
            "fieldData": {
                "part_number": "=NOT_A_FORMULA",
                "part_name_en": "=ALSO_TEXT",
                "status": "Inactive",
                "stock_on_hand_qty": "3.5",
                "safety_stock_qty": "",
            },
        },
    ])

    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Parts"]
    assert list(sheet.values) == [
        ("Part No.", "Part Name", "Status", "Inventory", "Safety Stock", "Turnover", "Created"),
        ("AL050013-00", "Pipe Holder", "Active", 8, 2, "12 Days", "12/05/2024"),
        ("=NOT_A_FORMULA", "=ALSO_TEXT", "Inactive", 3.5, 0, "0 Days", None),
    ]
    assert sheet["A3"].data_type == "s"
    assert sheet["B3"].data_type == "s"
    assert sheet.freeze_panes == "A2"


def test_product_excel_export_fetches_all_pages_with_customer_scope() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            rows = [
                {"fieldData": {"product_sku": "SKU-1", "stock": 7}},
                {"fieldData": {"product_sku": "SKU-2", "stock": 0}},
                {"fieldData": {"product_sku": "SKU-3", "stock": -2}},
            ]
            batch = rows[:2] if offset == 1 else rows[2:]
            return {"data": batch, "foundCount": 3, "returnedCount": len(batch)}

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )

    response = asyncio.run(export_customer_products(
        q="",
        session=session,
        filemaker=filemaker,
    ))

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["x-export-row-count"] == "3"
    assert response.headers["content-disposition"].endswith('.xlsx"')
    workbook = load_workbook(BytesIO(response.body), read_only=True)
    assert list(workbook["Products"].values) == [
        ("SKU", "Inventory"),
        ("SKU-1", 7),
        ("SKU-2", 0),
        ("SKU-3", -2),
    ]
    assert [call[3] for call in filemaker.calls] == [1, 3]
    assert all(call[0] == "@products" for call in filemaker.calls)
    assert all(call[1] == [{"id_client": "==CU638"}] for call in filemaker.calls)
    assert all(call[4] == [{"fieldName": "product_sku", "sortOrder": "ascend"}] for call in filemaker.calls)


def test_part_excel_export_fetches_all_search_results_with_customer_scope() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            rows = [
                {
                    "fieldData": {
                        "part_number": "AL050013-00",
                        "part_name_en": "Pipe Holder",
                        "stock_on_hand_qty": 8,
                        "safety_stock_qty": 2,
                    },
                },
                {
                    "fieldData": {
                        "part_number": "AL050013-01",
                        "part_name_en": "Pipe Holder B",
                        "stock_on_hand_qty": 4,
                        "safety_stock_qty": 1,
                    },
                },
                {
                    "fieldData": {
                        "part_number": "AL050013-02",
                        "part_name_en": "Pipe Holder C",
                        "stock_on_hand_qty": 0,
                        "safety_stock_qty": 0,
                    },
                },
            ]
            batch = rows[:2] if offset == 1 else rows[2:]
            return {"data": batch, "foundCount": 3, "returnedCount": len(batch)}

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )

    response = asyncio.run(export_customer_parts(
        q="AL050013",
        session=session,
        filemaker=filemaker,
    ))

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["x-export-row-count"] == "3"
    assert response.headers["content-disposition"].endswith('.xlsx"')
    workbook = load_workbook(BytesIO(response.body), read_only=True)
    assert list(workbook["Parts"].values) == [
        ("Part No.", "Part Name", "Status", "Inventory", "Safety Stock", "Turnover", "Created"),
        ("AL050013-00", "Pipe Holder", None, 8, 2, "0 Days", None),
        ("AL050013-01", "Pipe Holder B", None, 4, 1, "0 Days", None),
        ("AL050013-02", "Pipe Holder C", None, 0, 0, "0 Days", None),
    ]
    assert [call[3] for call in filemaker.calls] == [1, 3]
    assert all(call[0] == "Parts" for call in filemaker.calls)
    assert all(call[1] == [
        {"part_number": "*AL050013*", "customer_id": "==CU638"},
        {"part_name_en": "*AL050013*", "customer_id": "==CU638"},
    ] for call in filemaker.calls)
    assert all(call[4] == [{"fieldName": "part_number", "sortOrder": "ascend"}] for call in filemaker.calls)


def test_part_catalog_query_forces_customer_id_scope() -> None:
    query = _scoped_query(
        "AL05",
        search_fields=("part_number", PART_NAME_FIELD),
        scope_field=PART_SCOPE_FIELD,
        scope_value="CU638",
    )

    assert all(item["customer_id"] == "==CU638" for item in query)
    assert query[0]["part_number"] == "*AL05*"


def test_part_catalog_uses_current_filemaker_fields_for_scope_sort_and_rows() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {
                "data": [{
                    "recordId": "2",
                    "fieldData": {
                        "part_number": "AL050013-00",
                        "part_name_en": "Front arm",
                        "stock_on_hand_qty": 8,
                        "safety_stock_qty": 2,
                        "status": "Active",
                    },
                }],
                "foundCount": 1370,
                "returnedCount": 1,
            }

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )

    response = asyncio.run(list_customer_parts(
        q="AL05",
        page=1,
        page_size=10,
        sort_by="stock",
        sort_order="desc",
        session=session,
        filemaker=filemaker,
    ))

    assert response.found_count == 1370
    assert response.rows[0].part_name == "Front arm"
    assert response.rows[0].stock == 8
    assert response.rows[0].safety_stock == 2
    assert filemaker.calls == [(
        "Parts",
        [
            {"part_number": "*AL05*", "customer_id": "==CU638"},
            {"part_name_en": "*AL05*", "customer_id": "==CU638"},
        ],
        10,
        1,
        [{"fieldName": PART_STOCK_FIELD, "sortOrder": "descend"}],
    )]


def test_order_catalog_query_forces_web_client_scope_on_every_branch() -> None:
    query = _scoped_query(
        "UPS",
        search_fields=("shipping_company", "tracking_number"),
        scope_field=ORDER_SCOPE_FIELD,
        scope_value="0780",
    )

    assert query == [
        {
            "shipping_company": "*UPS*",
            "select_client_for_web_id": "==0780",
        },
        {
            "tracking_number": "*UPS*",
            "select_client_for_web_id": "==0780",
        },
    ]


def test_order_catalog_maps_only_requested_customer_visible_fields() -> None:
    order = _order(
        {
            "recordId": "91",
            "fieldData": {
                "出貨單 PI": "Sample order",
                "內部訂單單據編號": "NB001",
                "訂單 PO": "PO-7788",
                "shipping_company": "优速",
                "tracking_number": "910038198088",
                "order_remarks_for_client_only": "1252.12 x 7 = 8764.84",
                "log": "Hidden internal log",
            },
        },
        {
            "出貨單 PI": "Sample order",
            "內部訂單單據編號": "NB001",
            "訂單 PO": "PO-7788",
            "出貨單_客戶::客戶名稱": "欧先生",
            "貨款總和_price": 8764.84,
            "shipping_cost": 8.6,
            "出貨日期": "07/22/2026",
            "出货状态": "Shipped",
            "銀行費用": 99,
        },
    ).model_dump(by_alias=True)

    assert order == {
        "orderRef": "91",
        "clientName": "欧先生",
        "orderNumber": "PO-7788",
        "orderAmount": 8764.84,
        "shippingCompany": "优速",
        "trackingNumber": "910038198088",
        "shippingCost": 8.6,
        "shippedDate": "07/22/2026",
        "shippingStatus": "Shipped",
        "remarks": "1252.12 x 7 = 8764.84",
    }


def test_order_catalog_never_falls_back_to_internal_pi_or_internal_order_number() -> None:
    order = _order(
        {
            "recordId": "91",
            "fieldData": {
                "出貨單 PI": "PI-MY-USA-001",
                "內部訂單單據編號": "NB001",
                "訂單 PO": "",
            },
        },
        {
            "出貨單 PI": "PI-MY-USA-001",
            "內部訂單單據編號": "NB001",
            "訂單 PO": "",
            "出貨單_客戶::客戶名稱": "Mayako",
        },
    )

    payload = order.model_dump(by_alias=True)
    serialized = order.model_dump_json(by_alias=True)
    assert payload["orderNumber"] == ""
    assert "PI-MY-USA-001" not in serialized
    assert "NB001" not in serialized


def test_order_shipping_status_matches_the_shipped_date_filter() -> None:
    order = _order(
        {
            "recordId": "91",
            "fieldData": {
                "訂單 PO": "PO-7788",
                "出貨日期": "",
                "出货状态": "Shipped",
            },
        },
        {},
    )

    assert order.shipping_status == "Not Shipped"


def test_order_catalog_hides_legacy_pi_value_misfiled_as_customer_po() -> None:
    order = _order(
        {
            "recordId": "92",
            "fieldData": {
                "訂單 PO": "PI-MY-Australia-CID14-OID100003957",
            },
        },
        {},
    )

    assert order.order_number == ""
    assert "PI-MY-Australia" not in order.model_dump_json(by_alias=True)


def test_order_detail_lookup_uses_only_identities_from_scoped_primary_rows() -> None:
    class FakeFileMaker:
        async def find_records(self, layout, query=None, limit=100):
            assert layout == "@mayako"
            assert query == [{
                "select_client_for_web_id": "==0780",
                "內部訂單單據編號": "==NB001",
                "出貨單 PI": "==PI-001",
            }]
            assert limit == 2
            return {
                "data": [{
                    "fieldData": {
                        "內部訂單單據編號": "NB001",
                        "出貨單 PI": "PI-001",
                        "shipping_cost": 8.6,
                    },
                }],
                "foundCount": 1,
                "returnedCount": 1,
            }

    details = asyncio.run(_order_details(
        FakeFileMaker(),
        [{"fieldData": {"內部訂單單據編號": "NB001", "出貨單 PI": "PI-001"}}],
        "0780",
    ))

    assert details[("NB001", "PI-001")]["shipping_cost"] == 8.6


def test_order_endpoint_applies_session_web_client_scope_without_shipment_uuid() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {
                "data": [{
                    "recordId": "91",
                    "fieldData": {
                        "出貨單 PI": "PI-001",
                        "內部訂單單據編號": "NB001",
                        "出貨單_客戶::客戶名稱": "Mayako",
                        "shipping_company": "UPS",
                        "貨款總和_price": 1250,
                        "shipping_cost": 8.6,
                        "出貨日期": "07/22/2026",
                    },
                }],
                "foundCount": 1,
                "returnedCount": 1,
            }

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="",
        access_role="manager",
    )

    response = asyncio.run(list_customer_orders(
        q="UPS",
        page=1,
        page_size=10,
        sort_by="shippingCompany",
        sort_order="asc",
        session=session,
        filemaker=filemaker,
    ))

    primary_call = filemaker.calls[0]
    assert primary_call[0] == "@mayako"
    assert all(
        branch["select_client_for_web_id"]
        == "==0780"
        for branch in primary_call[1]
    )
    assert primary_call[4] == [{"fieldName": "shipping_company", "sortOrder": "ascend"}]
    assert response.rows[0].client_name == "Mayako"
    assert response.rows[0].order_amount == 1250
    assert response.rows[0].shipping_cost == 8.6
    assert response.rows[0].shipping_status == "Shipped"


def test_team_order_amounts_are_hidden_and_agent_order_access_is_blocked() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query))
            return {
                "data": [{
                    "recordId": "91",
                    "fieldData": {
                        "訂單 PO": "PO-001",
                        "貨款總和_price": 1250,
                        "shipping_cost": 8.6,
                    },
                }],
                "foundCount": 1,
                "returnedCount": 1,
            }

    def session(role: str) -> CustomerSession:
        return CustomerSession(
            session_id=f"{role}-session",
            username=role,
            display_name=role,
            client_name="Mayako",
            product_privilege="0780",
            part_customer_id="CU638",
            expires_at=9999999999,
            shipment_company_id="SHIP-1",
            access_role=role,
        )

    filemaker = FakeFileMaker()
    response = asyncio.run(list_customer_orders(
        q="",
        page=1,
        page_size=10,
        sort_by="orderNumber",
        sort_order="desc",
        session=session("team"),
        filemaker=filemaker,
    ))

    assert response.rows[0].order_amount is None
    assert response.rows[0].shipping_cost is None

    with pytest.raises(HTTPException) as exc_info:
        asyncio.run(list_customer_orders(
            q="",
            page=1,
            page_size=10,
            sort_by="orderNumber",
            sort_order="desc",
            session=session("agent"),
            filemaker=filemaker,
        ))
    assert exc_info.value.status_code == 403
    assert len(filemaker.calls) == 1


def test_order_month_and_shipping_status_filters_are_applied_to_every_search_branch() -> None:
    query = _order_catalog_query(
        "UPS",
        web_client_id="0780",
        month="2026-07",
        shipping_status="notShipped",
    )

    assert len(query) > 1
    assert all(
        branch[ORDER_SCOPE_FIELD] == "==0780"
        for branch in query
    )
    assert all(branch["訂單 PO"] in {"*", "*UPS*"} for branch in query)
    assert sum(branch["訂單 PO"] == "*UPS*" for branch in query) == 1
    assert all(branch["日期"] == "7/1/2026...7/31/2026" for branch in query)
    assert all(branch["出貨日期"] == "=" for branch in query)
    assert _order_month_range("2024-02") == "2/1/2024...2/29/2024"


def test_order_keyword_query_escapes_filemaker_find_operators_in_customer_po() -> None:
    query = _order_catalog_query(
        "PO#292687(CA1)",
        web_client_id="0780",
        month="",
        shipping_status="all",
    )

    assert any(branch["訂單 PO"] == r"*PO\#292687(CA1)*" for branch in query)
    assert all(branch[ORDER_SCOPE_FIELD].startswith("==") for branch in query)


def test_order_summary_uses_all_filtered_records_and_web_price_field() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            rows = [
                {"fieldData": {"貨款總和_price": "1,200.50", "出貨日期": "07/03/2026"}},
                {"fieldData": {"貨款總和_price": "50", "出貨日期": ""}},
                {"fieldData": {"貨款總和_price": "", "出貨日期": "", "出货状态": "TBC"}},
            ]
            return {"data": rows, "foundCount": 3, "returnedCount": 3}

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
        access_role="manager",
    )

    response = asyncio.run(summarize_customer_orders(
        q="",
        month="2026-07",
        shipping_status="all",
        session=session,
        filemaker=filemaker,
    ))

    assert response.order_amount_total == 1250.5
    assert response.order_count == 3
    assert response.shipped_count == 1
    assert response.not_shipped_count == 2
    assert filemaker.calls == [(
        "@mayako",
        [{
            ORDER_SCOPE_FIELD: "==0780",
            "訂單 PO": "*",
            "日期": "7/1/2026...7/31/2026",
        }],
        500,
        1,
        None,
    )]


def test_order_chat_date_and_hidden_field_search_scopes_every_branch() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {
                "data": [{
                    "recordId": "91",
                    "fieldData": {
                        "出貨單 PI": "PI-001",
                        "出貨單_客戶::客戶名稱": "Mayako",
                        "shipping_company": "UPS",
                        "tracking_number": "1Z999",
                        "shipping_cost": 8.6,
                        "出貨日期": "07/07/2026",
                    },
                }],
                "foundCount": 1,
                "returnedCount": 1,
            }

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )

    response = asyncio.run(find_customer_orders_for_chat(
        search="UPS",
        date_field="出貨日期",
        date_range="7/1/2026...7/22/2026",
        page=1,
        page_size=4,
        session=session,
        filemaker=filemaker,
    ))

    call = filemaker.calls[0]
    assert call[0] == "@mayako"
    assert len(call[1]) == len(ORDER_CHAT_TEXT_SEARCH_FIELDS)
    assert all(branch[ORDER_SCOPE_FIELD] == "==0780" for branch in call[1])
    assert all(branch["訂單 PO"] in {"*", "*UPS*"} for branch in call[1])
    assert sum(branch["訂單 PO"] == "*UPS*" for branch in call[1]) == 1
    assert all(branch["出貨日期"] == "7/1/2026...7/22/2026" for branch in call[1])
    assert any(branch.get("訂單 PO備註") == "*UPS*" for branch in call[1])
    assert call[4] == [{"fieldName": "訂單 PO", "sortOrder": "descend"}]
    assert response.rows[0].tracking_number == "1Z999"
    assert response.rows[0].shipped_date == "07/07/2026"


def test_order_chat_unshipped_search_uses_empty_shipped_date_and_scope() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {"data": [], "foundCount": 0, "returnedCount": 0}

    filemaker = FakeFileMaker()
    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )

    asyncio.run(find_customer_orders_for_chat(
        search="未出貨",
        date_field=None,
        date_range=None,
        page=1,
        page_size=4,
        session=session,
        filemaker=filemaker,
    ))

    query = filemaker.calls[0][1]
    assert query == [{
        ORDER_SCOPE_FIELD: "==0780",
        "訂單 PO": "*",
        "出貨日期": "=",
    }]
    assert filemaker.calls[0][4] == [{"fieldName": "訂單 PO", "sortOrder": "descend"}]


def test_order_chat_payment_date_uses_date_range_after_filemaker_conversion() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {"data": [], "foundCount": 0, "returnedCount": 0}

    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )

    filemaker = FakeFileMaker()
    response = asyncio.run(find_customer_orders_for_chat(
        search="",
        date_field="收款日期",
        date_range="7/1/2026...7/22/2026",
        page=1,
        page_size=4,
        session=session,
        filemaker=filemaker,
    ))

    assert response.found_count == 0
    assert filemaker.calls[0][1] == [{
        ORDER_SCOPE_FIELD: "==0780",
        "訂單 PO": "*",
        "收款日期": "7/1/2026...7/22/2026",
    }]
    assert filemaker.calls[0][4] == [{"fieldName": "訂單 PO", "sortOrder": "descend"}]


def test_order_chat_shipped_filter_uses_nonempty_shipped_date() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {"data": [], "foundCount": 0, "returnedCount": 0}

    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )
    filemaker = FakeFileMaker()

    asyncio.run(find_customer_orders_for_chat(
        search="",
        date_field=None,
        date_range=None,
        page=1,
        page_size=4,
        session=session,
        filemaker=filemaker,
        shipping_status="shipped",
    ))

    assert filemaker.calls[0][1] == [{
        ORDER_SCOPE_FIELD: "==0780",
        "訂單 PO": "*",
        "出貨日期": "*",
    }]


def test_order_chat_combines_keyword_order_month_and_not_shipped_filter() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {"data": [], "foundCount": 0, "returnedCount": 0}

    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )
    filemaker = FakeFileMaker()

    asyncio.run(find_customer_orders_for_chat(
        search="UPS",
        date_field="日期",
        date_range="7/1/2026...7/31/2026",
        page=1,
        page_size=4,
        session=session,
        filemaker=filemaker,
        shipping_status="notShipped",
    ))

    query = filemaker.calls[0][1]
    assert len(query) == len(ORDER_CHAT_TEXT_SEARCH_FIELDS)
    assert all(
        branch[ORDER_SCOPE_FIELD] == "==0780"
        for branch in query
    )
    assert all(branch["訂單 PO"] in {"*", "*UPS*"} for branch in query)
    assert all(branch["日期"] == "7/1/2026...7/31/2026" for branch in query)
    assert all(branch["出貨日期"] == "=" for branch in query)


def test_order_chat_broad_listing_counts_and_pages_only_customer_po_records() -> None:
    class FakeFileMaker:
        def __init__(self) -> None:
            self.calls = []

        async def find_records(self, layout, query=None, limit=100, offset=1, sort=None):
            self.calls.append((layout, query, limit, offset, sort))
            return {
                "data": [
                    {"recordId": "1", "fieldData": {"訂單 PO": "PO#1001"}},
                    {"recordId": "2", "fieldData": {"訂單 PO": "100005958"}},
                ],
                "foundCount": 5,
                "returnedCount": 2,
            }

    session = CustomerSession(
        session_id="session",
        username="mayako",
        display_name="Mayako",
        client_name="Mayako",
        product_privilege="0780",
        part_customer_id="CU638",
        expires_at=9999999999,
        shipment_company_id="0E254109-8698-4F5D-BE70-ABFD2B929CE9",
    )
    filemaker = FakeFileMaker()

    response = asyncio.run(find_customer_orders_for_chat(
        search="",
        date_field=None,
        date_range=None,
        page=1,
        page_size=2,
        session=session,
        filemaker=filemaker,
    ))

    assert filemaker.calls[0][1] == [{
        ORDER_SCOPE_FIELD: "==0780",
        "訂單 PO": "*",
    }]
    assert response.found_count == 5
    assert response.returned_count == 2
    assert response.total_pages == 3
    assert [row.order_number for row in response.rows] == ["PO#1001", "100005958"]


def test_catalog_rows_only_map_customer_safe_fields() -> None:
    product = _product(
        {
            "recordId": "1",
            "fieldData": {
                "product_sku": "P-1",
                "product_name": "Product",
                "stock": 4,
                "BOM計數": 3,
                "產品售價::Price": 99,
                "產品 BOM::廠商": "Hidden vendor",
            },
        }
    ).model_dump(by_alias=True)
    part = _part(
        {
            "recordId": "2",
            "fieldData": {
                "part_number": "A-1",
                "part_name": "Chinese-only internal name",
                "part_name_en": "Part",
                "stock_on_hand_qty": 8,
                "safety_stock_qty": 2,
                "Turnover Time": "12 Days",
                "Date Created": "12/05/2024",
                "cost": 12,
            },
        }
    ).model_dump(by_alias=True)

    assert product == {
        "productRef": "1",
        "productSku": "P-1",
        "productName": "Product",
        "modelName": "",
        "scale": "",
        "category": "",
        "stock": 4,
        "bomCount": 3,
        "hasImage": False,
    }
    assert part == {
        "partRef": "2",
        "partNumber": "A-1",
        "partName": "Part",
        "stock": 8,
        "safetyStock": 2,
        "turnover": "12 Days",
        "created": "12/05/2024",
        "status": "",
        "hasImage": False,
    }


def test_product_detail_maps_requested_inventory_metrics_only() -> None:
    detail = _product_detail({
        "recordId": "9",
        "fieldData": {
            "product_sku": "MYB0196",
            "stock": 7,
            "產品庫存::出庫數量總合": 20,
            "產品售價::Price": 1.9,
            "Stock_USD": 13.3,
            "PrePaid_stock_USD": 13.3,
            "下單數量": 103,
            "cost": 0.5,
            "vendor": "Hidden supplier",
        },
    }).model_dump(by_alias=True)

    assert detail["stock"] == 7
    assert detail["soldTotal"] == 20
    assert detail["price"] == 1.9
    assert detail["stockValue"] == 13.3
    assert detail["prepaidStock"] == 13.3
    assert detail["productionCalculation"] == 103
    assert "cost" not in detail and "vendor" not in detail


def test_product_detail_omits_price_when_account_cannot_view_it() -> None:
    detail = _product_detail(
        {
            "recordId": "9",
            "fieldData": {
            "product_sku": "MYB0196",
            "stock": 7,
            "產品售價::Price": 1.9,
            "Stock_USD": 13.3,
            "PrePaid_stock_USD": 13.3,
            },
        },
        can_view_price=False,
    ).model_dump(by_alias=True, exclude_unset=True)

    assert "price" not in detail
    assert "stockValue" not in detail
    assert "prepaidStock" not in detail


def test_product_images_are_exactly_scoped_to_migrated_source_record() -> None:
    class FakeFileMaker:
        async def find_records(self, layout, query=None, limit=100, sort=None):
            assert layout == "ProductAssets"
            assert query == {
                "source_record_id": "==15572",
                "id_client_snapshot": "==CU638",
                "asset_type": "==product_image",
                "visibility": "==customer",
                "migration_status": "==copied",
            }
            assert limit == 100
            assert sort == [{"fieldName": "sort_order", "sortOrder": "ascend"}]
            return {"data": [], "foundCount": 0, "returnedCount": 0}

    result = asyncio.run(_product_images(FakeFileMaker(), "15572", "CU638"))

    assert result["data"] == []


def test_product_image_metadata_exposes_no_container_url() -> None:
    image = _product_image({
        "recordId": "13612",
        "fieldData": {
            "asset_file": "https://filemaker.example/container/secret",
            "original_filename": "side-view.jpg",
            "title": "Side view",
            "sort_order": 2,
            "is_primary": 0,
        },
    }).model_dump(by_alias=True)

    assert image == {
        "assetRef": "13612",
        "filename": "side-view.jpg",
        "title": "Side view",
        "sortOrder": 2,
        "isPrimary": False,
    }
    assert "asset_file" not in image


def test_related_products_are_reverse_linked_scoped_and_deduplicated() -> None:
    class FakeFileMaker:
        async def find_records(self, layout, query=None, limit=100, sort=None):
            if layout == "@product_bom":
                assert query == {"零件編號": "==A-1"}
                return {
                    "data": [
                        {"fieldData": {"ID_產品編號": "P-2"}},
                        {"fieldData": {"ID_產品編號": "P-1"}},
                        {"fieldData": {"ID_產品編號": "P-1"}},
                    ]
                }
            assert layout == "@products"
            assert all(item["id_client"] == "==CU638" for item in query)
            return {
                "data": [
                    {"recordId": "10", "fieldData": {"product_sku": "P-1", "product_name": "One", "產品售價::Price": 99}},
                    {"recordId": "20", "fieldData": {"product_sku": "P-2", "product_name": "Two"}},
                ]
            }

    products = asyncio.run(_related_products(FakeFileMaker(), "A-1", "CU638"))

    assert [product.model_dump(by_alias=True) for product in products] == [
        {"productRef": "10", "productSku": "P-1", "productName": "One"},
        {"productRef": "20", "productSku": "P-2", "productName": "Two"},
    ]


def test_customer_bom_lines_are_composed_from_api_layout_parts() -> None:
    class FakeFileMaker:
        async def find_records(self, layout, query=None, limit=100):
            assert layout == "@零件"
            assert query == [{"part_number": "==A-1"}]
            return {
                "data": [{
                    "recordId": "7",
                    "fieldData": {
                        "part_number": "A-1",
                        "part_name_en": "Front arm",
                        "stock_on_hand_qty": 8,
                        "status": "Active",
                    },
                }],
                "foundCount": 1,
                "returnedCount": 1,
            }

    lines = asyncio.run(_customer_bom_lines(FakeFileMaker(), {
        "data": [{
            "recordId": "21",
            "fieldData": {
                "零件編號": "A-1",
                "需求數量": 2,
                "倉庫需求": 4,
            },
        }],
    }))

    assert [line.model_dump(by_alias=True) for line in lines] == [{
        "lineRef": "21",
        "partNumber": "A-1",
        "clientPartNumber": "",
        "partName": "Front arm",
        "bomQuantity": 2,
        "requiredQuantity": 4,
        "stock": 8,
        "status": "Active",
        "sparePartNumber": "",
        "spareStock": None,
    }]


def test_internal_statuses_are_translated_or_hidden() -> None:
    assert _english_status("待确认") == "Pending confirmation"
    assert _english_status("Active") == "Active"
    assert _english_status("未知状态") == ""
    assert _english_text("零件包") == "Parts kit"
    assert _english_text("Chinese chassis 车子底盘") == ""
    assert _english_text("MPC マヤコ Mayako MX8") == "MPC Mayako MX8"


def test_part_blank_stock_metrics_match_filemaker_display_defaults() -> None:
    part = _part({"recordId": "2", "fieldData": {"part_number": "A-2"}})

    assert part.safety_stock == 0
    assert part.turnover == "0 Days"
