import asyncio
import re
from calendar import monthrange
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from math import ceil
from typing import Annotated, Any, Literal

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models.customer_catalog import (
    CustomerBomLine,
    CustomerCatalogOrder,
    CustomerCatalogPart,
    CustomerCatalogProduct,
    CustomerOrderListResponse,
    CustomerOrderSummaryResponse,
    CustomerPartDetailResponse,
    CustomerPartListResponse,
    CustomerProductImage,
    CustomerProductDetailResponse,
    CustomerProductDetailItem,
    CustomerProductListResponse,
    CustomerRelatedProduct,
)
from app.services.customer_chat_auth import CustomerSession
from app.services.dependencies import get_customer_session, get_filemaker_client
from app.services.filemaker_client import FileMakerAPIError, FileMakerClient
from app.services.product_api import (
    PRODUCT_ASSET_LAYOUT,
    PRODUCT_BOM_LAYOUT,
    PRODUCT_LAYOUT,
    PRODUCT_PART_LAYOUT,
    PRODUCT_STOCK_FIELD,
    find_product_bom,
    find_product_price,
    price_value,
)


router = APIRouter(prefix="/customer-chat/catalog", tags=["customer-catalog"])
PART_LAYOUT = "Parts"
PART_SCOPE_FIELD = "customer_id"
PART_NAME_FIELD = "part_name_en"
PART_STOCK_FIELD = "stock_on_hand_qty"
PART_SAFETY_STOCK_FIELD = "safety_stock_qty"
ORDER_LAYOUT = "@出貨單"
# This is an external FileMaker layout identifier, not the portal brand. Keep it
# until the FileMaker layout is migrated separately.
LEGACY_ORDER_DETAIL_LAYOUT = "@mayako"
ORDER_SCOPE_FIELD = "select_client_for_web_id"
ORDER_ID_FIELD = "id"
ORDER_INTERNAL_ID_FIELD = "internal_id"
ORDER_AMOUNT_FIELD = "貨款總和_price"
MAX_CATALOG_PAGE_SIZE = 100
PRODUCT_EXPORT_PAGE_SIZE = 500
MAX_PRODUCT_EXPORT_ROWS = 10_000
PART_EXPORT_PAGE_SIZE = 500
MAX_PART_EXPORT_ROWS = 10_000
ORDER_SUMMARY_PAGE_SIZE = 500
MAX_ORDER_SUMMARY_ROWS = 20_000

PRODUCT_SEARCH_FIELDS = ("product_sku", "系統產品編號", "product_name", "車款", "類別")
PART_SEARCH_FIELDS = ("part_number", PART_NAME_FIELD)
ORDER_SEARCH_FIELDS = (
    ORDER_ID_FIELD,
    "出貨單 PI",
    "包裝狀態",
    "客戶備註",
    "訂單 PO",
    ORDER_INTERNAL_ID_FIELD,
    "shipping_company",
    "tracking_number",
    "order_remarks_for_client_only",
    "shipping_notes",
)
ORDER_CHAT_TEXT_SEARCH_FIELDS = (
    ORDER_ID_FIELD,
    "出貨單 PI",
    ORDER_INTERNAL_ID_FIELD,
    "訂單 PO",
    "訂單 PO備註",
    "出货状态",
    "包裝狀態",
    "付款狀態",
    "訂單分類",
    "訂單型態",
    "概要",
    "order_status",
    "訂單概要中文",
    "shipping_company",
    "快遞出貨公司名稱",
    "快遞出貨國家",
    "tracking_number",
    "order_remarks_for_client_only",
    "shipping_notes",
    "客戶備註",
    "出貨單_客戶::客戶備註",
    "出貨單_客戶::客戶名稱",
    "出貨單_客戶::客戶代號",
    "出貨單_客戶::客戶公司簡稱",
    "出貨單資料::客戶SKU",
    "出貨單資料::產品名稱",
    "出貨單資料::產品編號",
)
ORDER_CHAT_NUMBER_SEARCH_FIELDS = (
    "shipping_cost",
    ORDER_AMOUNT_FIELD,
)
ORDER_CHAT_DATE_FIELDS = frozenset({
    "日期",
    "備好日期",
    "出貨日期",
    "完成日期",
    "收款日期",
    "簽名日期",
    "updated_at",
})
PRODUCT_SORT_FIELDS = {
    "productSku": "product_sku",
    "productName": "product_name",
    "modelName": "車款",
    "scale": "車子比例",
    "category": "類別",
    "stock": PRODUCT_STOCK_FIELD,
    "bomCount": "BOM計數",
}
PART_SORT_FIELDS = {
    "partNumber": "part_number",
    "partName": PART_NAME_FIELD,
    "stock": PART_STOCK_FIELD,
    "status": "status",
}
ORDER_SORT_FIELDS = {
    "orderNumber": "訂單 PO",
    "orderAmount": ORDER_AMOUNT_FIELD,
    "shippingCompany": "shipping_company",
    "trackingNumber": "tracking_number",
    "shippedDate": "出貨日期",
}


@router.get("/products", response_model=CustomerProductListResponse)
async def list_customer_products(
    q: str = Query(default="", max_length=80),
    page: int = Query(default=1, ge=1, le=10_000),
    page_size: int = Query(default=10, ge=1, le=MAX_CATALOG_PAGE_SIZE, alias="pageSize"),
    sort_by: str = Query(default="productSku", alias="sortBy"),
    sort_order: Literal["asc", "desc"] = Query(default="asc", alias="sortOrder"),
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> CustomerProductListResponse:
    normalized_query = q.strip()
    selected_sort = sort_by if sort_by in PRODUCT_SORT_FIELDS else "productSku"
    try:
        result = await filemaker.find_records(
            PRODUCT_LAYOUT,
            query=_scoped_query(
                normalized_query,
                search_fields=PRODUCT_SEARCH_FIELDS,
                scope_field="id_client",
                scope_value=session.part_customer_id,
            ),
            limit=page_size,
            offset=((page - 1) * page_size) + 1,
            sort=[{
                "fieldName": PRODUCT_SORT_FIELDS[selected_sort],
                "sortOrder": "ascend" if sort_order == "asc" else "descend",
            }],
        )
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc
    found_count = int(result["foundCount"] or 0)
    return CustomerProductListResponse(
        rows=[
            _product(record, inventory_only=not session.can_view_details)
            for record in result["data"]
        ],
        foundCount=found_count,
        returnedCount=int(result["returnedCount"] or len(result["data"])),
        page=page,
        pageSize=page_size,
        totalPages=max(1, ceil(found_count / page_size)),
        query=normalized_query,
        sortBy=selected_sort,
        sortOrder=sort_order,
    )


@router.get("/products/export.xlsx")
async def export_customer_products(
    q: str = Query(default="", max_length=80),
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> Response:
    """Export every scoped product SKU, bilingual name, and inventory value."""
    normalized_query = q.strip()
    query = _scoped_query(
        normalized_query,
        search_fields=PRODUCT_SEARCH_FIELDS,
        scope_field="id_client",
        scope_value=session.part_customer_id,
    )
    records: list[dict[str, Any]] = []
    found_count: int | None = None
    try:
        while found_count is None or len(records) < found_count:
            result = await filemaker.find_records(
                PRODUCT_LAYOUT,
                query=query,
                limit=PRODUCT_EXPORT_PAGE_SIZE,
                offset=len(records) + 1,
                sort=[{"fieldName": "product_sku", "sortOrder": "ascend"}],
            )
            found_count = int(result["foundCount"] or 0)
            if found_count > MAX_PRODUCT_EXPORT_ROWS:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={"message": "The export contains too many products. Narrow the search and try again."},
                )
            batch = list(result["data"])
            records.extend(batch)
            if not batch:
                break
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc

    content = _product_inventory_workbook(records)
    filename = f"stock-check-products-inventory-{date.today().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Row-Count": str(len(records)),
        },
    )


@router.get("/parts", response_model=CustomerPartListResponse)
async def list_customer_parts(
    q: str = Query(default="", max_length=80),
    page: int = Query(default=1, ge=1, le=10_000),
    page_size: int = Query(default=10, ge=1, le=MAX_CATALOG_PAGE_SIZE, alias="pageSize"),
    sort_by: str = Query(default="partNumber", alias="sortBy"),
    sort_order: Literal["asc", "desc"] = Query(default="asc", alias="sortOrder"),
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> CustomerPartListResponse:
    normalized_query = q.strip()
    selected_sort = sort_by if sort_by in PART_SORT_FIELDS else "partNumber"
    try:
        result = await filemaker.find_records(
            PART_LAYOUT,
            query=_scoped_query(
                normalized_query,
                search_fields=PART_SEARCH_FIELDS,
                scope_field=PART_SCOPE_FIELD,
                scope_value=session.part_customer_id,
            ),
            limit=page_size,
            offset=((page - 1) * page_size) + 1,
            sort=[{
                "fieldName": PART_SORT_FIELDS[selected_sort],
                "sortOrder": "ascend" if sort_order == "asc" else "descend",
            }],
        )
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc
    found_count = int(result["foundCount"] or 0)
    return CustomerPartListResponse(
        rows=[
            _part(record, inventory_only=not session.can_view_details)
            for record in result["data"]
        ],
        foundCount=found_count,
        returnedCount=int(result["returnedCount"] or len(result["data"])),
        page=page,
        pageSize=page_size,
        totalPages=max(1, ceil(found_count / page_size)),
        query=normalized_query,
        sortBy=selected_sort,
        sortOrder=sort_order,
    )


@router.get("/parts/export.xlsx")
async def export_customer_parts(
    q: str = Query(default="", max_length=80),
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> Response:
    """Export all customer-scoped parts and their visible inventory fields."""
    normalized_query = q.strip()
    query = _scoped_query(
        normalized_query,
        search_fields=PART_SEARCH_FIELDS,
        scope_field=PART_SCOPE_FIELD,
        scope_value=session.part_customer_id,
    )
    records: list[dict[str, Any]] = []
    found_count: int | None = None
    try:
        while found_count is None or len(records) < found_count:
            result = await filemaker.find_records(
                PART_LAYOUT,
                query=query,
                limit=PART_EXPORT_PAGE_SIZE,
                offset=len(records) + 1,
                sort=[{"fieldName": "part_number", "sortOrder": "ascend"}],
            )
            found_count = int(result["foundCount"] or 0)
            if found_count > MAX_PART_EXPORT_ROWS:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={"message": "The export contains too many parts. Narrow the search and try again."},
                )
            batch = list(result["data"])
            records.extend(batch)
            if not batch:
                break
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc

    content = _part_inventory_workbook(records)
    filename = f"stock-check-parts-inventory-{date.today().isoformat()}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Export-Row-Count": str(len(records)),
        },
    )


@router.get("/orders", response_model=CustomerOrderListResponse)
async def list_customer_orders(
    q: str = Query(default="", max_length=80),
    page: int = Query(default=1, ge=1, le=10_000),
    page_size: int = Query(default=10, ge=1, le=MAX_CATALOG_PAGE_SIZE, alias="pageSize"),
    sort_by: str = Query(default="orderNumber", alias="sortBy"),
    sort_order: Literal["asc", "desc"] = Query(default="desc", alias="sortOrder"),
    month: Annotated[
        str,
        Query(pattern=r"^(?:\d{4}-(?:0[1-9]|1[0-2]))?$"),
    ] = "",
    shipping_status: Annotated[
        Literal["all", "shipped", "notShipped"],
        Query(alias="shippingStatus"),
    ] = "all",
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> CustomerOrderListResponse:
    _require_order_access(session)
    return await find_customer_orders(
        q=q,
        page=page,
        page_size=page_size,
        sort_by=sort_by,
        sort_order=sort_order,
        session=session,
        filemaker=filemaker,
        month=month,
        shipping_status=shipping_status,
    )


@router.get("/orders/summary", response_model=CustomerOrderSummaryResponse)
async def summarize_customer_orders(
    q: str = Query(default="", max_length=80),
    month: Annotated[
        str,
        Query(pattern=r"^(?:\d{4}-(?:0[1-9]|1[0-2]))?$"),
    ] = "",
    shipping_status: Annotated[
        Literal["all", "shipped", "notShipped"],
        Query(alias="shippingStatus"),
    ] = "all",
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> CustomerOrderSummaryResponse:
    _require_order_access(session)
    web_client_id = _order_web_client_id(session)
    normalized_query = q.strip()
    query = _order_catalog_query(
        normalized_query,
        web_client_id=web_client_id,
        month=month,
        shipping_status=shipping_status,
    )
    try:
        records = await _all_order_records(filemaker, query)
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc
    return _order_summary(
        records,
        month=month,
        shipping_status=shipping_status,
        can_view_price=session.can_view_price,
    )


async def find_customer_orders(
    *,
    q: str,
    page: int,
    page_size: int,
    sort_by: str,
    sort_order: Literal["asc", "desc"],
    session: CustomerSession,
    filemaker: FileMakerClient,
    month: str = "",
    shipping_status: Literal["all", "shipped", "notShipped"] = "all",
) -> CustomerOrderListResponse:
    """Return customer-visible shipment records for catalog and chat callers."""
    _require_order_access(session)
    web_client_id = _order_web_client_id(session)

    normalized_query = q.strip()
    selected_sort = sort_by if sort_by in ORDER_SORT_FIELDS else "orderNumber"
    query = _order_catalog_query(
        normalized_query,
        web_client_id=web_client_id,
        month=month,
        shipping_status=shipping_status,
    )
    try:
        result = await filemaker.find_records(
            LEGACY_ORDER_DETAIL_LAYOUT,
            query=query,
            limit=page_size,
            offset=((page - 1) * page_size) + 1,
            sort=[{
                "fieldName": ORDER_SORT_FIELDS[selected_sort],
                "sortOrder": "ascend" if sort_order == "asc" else "descend",
            }],
        )
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc

    found_count = int(result["foundCount"] or 0)
    return CustomerOrderListResponse(
        rows=[
            _order(
                record,
                _fields(record),
                can_view_price=session.can_view_price,
            )
            for record in result["data"]
        ],
        foundCount=found_count,
        returnedCount=int(result["returnedCount"] or len(result["data"])),
        page=page,
        pageSize=page_size,
        totalPages=max(1, ceil(found_count / page_size)),
        query=normalized_query,
        sortBy=selected_sort,
        sortOrder=sort_order,
    )


async def find_customer_orders_for_chat(
    *,
    search: str,
    date_field: str | None,
    date_range: str | None,
    page: int,
    page_size: int,
    session: CustomerSession,
    filemaker: FileMakerClient,
    shipping_status: Literal["all", "shipped", "notShipped"] = "all",
) -> CustomerOrderListResponse:
    """Search the richer order layout while preserving the strict public boundary."""
    _require_order_access(session)
    web_client_id = _order_web_client_id(session)
    if date_field is not None and date_field not in ORDER_CHAT_DATE_FIELDS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"message": "That order date field is not available."},
        )

    normalized_search = search.strip()[:80]
    scope = f"=={web_client_id}"
    base: dict[str, str] = {
        ORDER_SCOPE_FIELD: scope,
        "訂單 PO": "*",
    }
    if date_field and date_range:
        base[date_field] = date_range

    effective_shipping_status = shipping_status
    if normalized_search in {"未出貨", "Not Shipped"} and shipping_status == "all":
        normalized_search = ""
        effective_shipping_status = "notShipped"
    elif normalized_search in {"已出貨", "Shipped"} and shipping_status == "all":
        normalized_search = ""
        effective_shipping_status = "shipped"

    if effective_shipping_status == "shipped" and date_field != "出貨日期":
        base["出貨日期"] = "*"
    elif effective_shipping_status == "notShipped" and date_field != "出貨日期":
        base["出貨日期"] = "="

    if (
        effective_shipping_status == "notShipped"
        and date_field == "出貨日期"
        and date_range
    ):
        return CustomerOrderListResponse(
            rows=[],
            foundCount=0,
            returnedCount=0,
            page=page,
            pageSize=page_size,
            totalPages=1,
            query=normalized_search,
            sortBy="orderNumber",
            sortOrder="desc",
        )

    query: list[dict[str, str]]
    if normalized_search:
        text_search_fields = [field for field in ORDER_CHAT_TEXT_SEARCH_FIELDS if field != date_field]
        query = [
            {**base, field: _filemaker_contains(normalized_search)}
            for field in text_search_fields
        ]
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", normalized_search.replace(",", "")):
            numeric = normalized_search.replace(",", "")
            query.extend(
                {**base, field: f"=={numeric}"}
                for field in ORDER_CHAT_NUMBER_SEARCH_FIELDS
            )
    else:
        query = [base]

    try:
        result = await filemaker.find_records(
            LEGACY_ORDER_DETAIL_LAYOUT,
            query=query,
            limit=page_size,
            offset=((page - 1) * page_size) + 1,
            sort=[{
                "fieldName": "訂單 PO",
                "sortOrder": "descend",
            }],
        )
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc

    found_count = int(result["foundCount"] or 0)
    return CustomerOrderListResponse(
        rows=[
            _order(record, _fields(record), can_view_price=session.can_view_price)
            for record in result["data"]
        ],
        foundCount=found_count,
        returnedCount=int(result["returnedCount"] or len(result["data"])),
        page=page,
        pageSize=page_size,
        totalPages=max(1, ceil(found_count / page_size)),
        query=normalized_search,
        sortBy="orderNumber",
        sortOrder="desc",
    )


@router.get(
    "/products/{record_id}",
    response_model=CustomerProductDetailResponse,
    response_model_exclude_unset=True,
)
async def get_customer_product_detail(
    record_id: str,
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> CustomerProductDetailResponse:
    _require_detail_access(session)
    record = await _get_record(filemaker, PRODUCT_LAYOUT, record_id)
    fields = _fields(record)
    if _text(fields.get("id_client")) != session.part_customer_id:
        raise _not_found("Product")
    product_sku = _text(fields.get("product_sku"))
    system_product_sku = _text(fields.get("系統產品編號"))
    try:
        detail_results = await asyncio.gather(
            find_product_bom(filemaker, product_sku),
            _product_images(filemaker, record_id, session.part_customer_id),
            *(
                [find_product_price(filemaker, product_sku, system_product_sku)]
                if session.can_view_price
                else []
            ),
        )
        bom_result, image_result = detail_results[:2]
        price_result = detail_results[2] if session.can_view_price else None
        bom = await _customer_bom_lines(filemaker, bom_result)
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc
    bom_found_count = int(bom_result.get("foundCount") or len(bom))
    bom_truncated = bom_found_count > len(bom)
    warnings: list[str] = []
    if bom_truncated:
        warnings.append(
            f"This product has {bom_found_count} BOM lines; only the first {len(bom)} are shown."
        )
    if "產品庫存::出庫數量總合" not in _fields(record):
        warnings.append("Sold total is not available from the API product layouts.")
    images = [_product_image(item) for item in image_result["data"]]
    return CustomerProductDetailResponse(
        product=_product_detail(
            record,
            can_view_price=session.can_view_price,
            price=(
                price_value(price_result, product_sku, system_product_sku)
                if price_result is not None
                else None
            ),
        ),
        images=images,
        imageCount=len(images),
        bom=bom,
        bomFoundCount=bom_found_count,
        bomReturnedCount=len(bom),
        bomTruncated=bom_truncated,
        warnings=warnings,
    )


async def _product_images(
    filemaker: FileMakerClient,
    source_record_id: str,
    client_id: str,
) -> dict[str, Any]:
    """Read customer-visible product images copied into ProductAssets.

    The immutable migration source record ID keeps duplicate historical product
    numbers from sharing images in the customer portal.
    """
    if not source_record_id.isdigit() or not client_id.strip():
        return {"data": [], "foundCount": 0, "returnedCount": 0}
    return await filemaker.find_records(
        PRODUCT_ASSET_LAYOUT,
        query={
            "source_record_id": f"=={source_record_id}",
            "id_client_snapshot": f"=={client_id.strip()}",
            "asset_type": "==product_image",
            "visibility": "==customer",
            "migration_status": "==copied",
        },
        limit=100,
        sort=[{"fieldName": "sort_order", "sortOrder": "ascend"}],
    )


@router.get("/parts/{record_id}", response_model=CustomerPartDetailResponse)
async def get_customer_part_detail(
    record_id: str,
    session: CustomerSession = Depends(get_customer_session),
    filemaker: FileMakerClient = Depends(get_filemaker_client),
) -> CustomerPartDetailResponse:
    _require_detail_access(session)
    record = await _get_record(filemaker, PART_LAYOUT, record_id)
    if not await _part_record_in_scope(filemaker, record, session.part_customer_id):
        raise _not_found("Part")
    part = _part(record)
    try:
        related_products = await _related_products(
            filemaker,
            part.part_number,
            session.part_customer_id,
        )
    except FileMakerAPIError as exc:
        raise _catalog_unavailable() from exc
    return CustomerPartDetailResponse(part=part, relatedProducts=related_products)


def _scoped_query(
    value: str,
    *,
    search_fields: tuple[str, ...],
    scope_field: str,
    scope_value: str,
) -> list[dict[str, str]]:
    scope = f"=={scope_value.strip()}"
    if not value:
        return [{scope_field: scope}]
    return [
        {
            field: f"*{value}*",
            scope_field: scope,
        }
        for field in search_fields
    ]


def _order_web_client_id(session: CustomerSession) -> str:
    web_client_id = session.product_privilege.strip()
    if not web_client_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={"message": "Order access is not configured for this account."},
        )
    return web_client_id


def _order_catalog_query(
    value: str,
    *,
    web_client_id: str,
    month: str,
    shipping_status: Literal["all", "shipped", "notShipped"],
) -> list[dict[str, str]]:
    base = {
        ORDER_SCOPE_FIELD: f"=={web_client_id.strip()}",
        "訂單 PO": "*",
    }
    date_range = _order_month_range(month)
    if date_range:
        base["日期"] = date_range
    if shipping_status == "shipped":
        base["出貨日期"] = "*"
    elif shipping_status == "notShipped":
        base["出貨日期"] = "="

    if not value:
        return [base]
    return [
        {**base, field: _filemaker_contains(value)}
        for field in ORDER_SEARCH_FIELDS
    ]


def _filemaker_contains(value: str) -> str:
    """Build a contains query without treating customer text as Find operators."""
    escaped = re.sub(r'([\\@*#?!="<>])', r"\\\1", value)
    return f"*{escaped}*"


def _order_month_range(month: str) -> str:
    normalized = month.strip()
    if not normalized:
        return ""
    match = re.fullmatch(r"(\d{4})-(0[1-9]|1[0-2])", normalized)
    if not match:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail={"message": "Order month must use YYYY-MM format."},
        )
    year = int(match.group(1))
    month_number = int(match.group(2))
    final_day = monthrange(year, month_number)[1]
    return f"{month_number}/1/{year}...{month_number}/{final_day}/{year}"


async def _all_order_records(
    filemaker: FileMakerClient,
    query: list[dict[str, str]],
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    found_count: int | None = None
    while found_count is None or len(records) < found_count:
        result = await filemaker.find_records(
            LEGACY_ORDER_DETAIL_LAYOUT,
            query=query,
            limit=ORDER_SUMMARY_PAGE_SIZE,
            offset=len(records) + 1,
        )
        found_count = int(result["foundCount"] or 0)
        if found_count > MAX_ORDER_SUMMARY_ROWS:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={"message": "There are too many orders to summarize. Select a month first."},
            )
        batch = [record for record in result["data"] if isinstance(record, dict)]
        records.extend(batch)
        if not batch:
            break
    return records


def _order_summary(
    records: list[dict[str, Any]],
    *,
    month: str,
    shipping_status: Literal["all", "shipped", "notShipped"],
    can_view_price: bool = True,
) -> CustomerOrderSummaryResponse:
    order_amount_total = Decimal()
    shipped_count = 0
    not_shipped_count = 0
    for record in records:
        fields = _fields(record)
        order_amount_total += _money_value(fields.get(ORDER_AMOUNT_FIELD))
        if _shipping_status(fields) == "Shipped":
            shipped_count += 1
        else:
            not_shipped_count += 1
    return CustomerOrderSummaryResponse(
        orderAmountTotal=float(order_amount_total) if can_view_price else None,
        orderCount=len(records),
        shippedCount=shipped_count,
        notShippedCount=not_shipped_count,
        month=month,
        shippingStatus=shipping_status,
    )


def _money_value(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal()
    normalized = str(value).strip().replace(",", "").replace("$", "")
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return Decimal()


def _require_order_access(session: CustomerSession) -> None:
    if not session.can_view_orders:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "message": "Your account does not have permission to view orders.",
                "code": "order_permission",
            },
        )


def _require_detail_access(session: CustomerSession) -> None:
    if not session.can_view_details:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "message": "This account is limited to inventory lookup.",
                "code": "inventory_only",
            },
        )


async def _get_record(filemaker: FileMakerClient, layout: str, record_id: str) -> dict[str, Any]:
    if not record_id.isdigit():
        raise _not_found("Record")
    try:
        data = await filemaker.get_record(layout, record_id)
    except FileMakerAPIError as exc:
        raise _not_found("Record") from exc
    record = data[0] if isinstance(data, list) and data else data if isinstance(data, dict) else None
    if not isinstance(record, dict):
        raise _not_found("Record")
    return record


async def _part_record_in_scope(
    filemaker: FileMakerClient,
    record: dict[str, Any],
    part_customer_id: str,
) -> bool:
    part_number = _text(_fields(record).get("part_number"))
    if not part_number:
        return False
    try:
        result = await filemaker.find_records(
            PART_LAYOUT,
            query={"part_number": f"=={part_number}", PART_SCOPE_FIELD: f"=={part_customer_id}"},
            limit=100,
        )
    except FileMakerAPIError:
        return False
    record_id = str(record.get("recordId") or "")
    return any(str(item.get("recordId") or "") == record_id for item in result["data"])


async def _customer_bom_lines(
    filemaker: FileMakerClient,
    bom_result: dict[str, Any],
) -> list[CustomerBomLine]:
    """Map a verified customer's product BOM using API-only layouts.

    Product ownership is checked before this function is called. Part lookups are
    restricted to exact numbers already present in that authorized BOM.
    """
    bom_records = [item for item in bom_result.get("data", []) if isinstance(item, dict)]
    part_numbers = list(dict.fromkeys(
        _text(_fields(record).get("零件編號"))
        for record in bom_records
        if _text(_fields(record).get("零件編號"))
    ))[:500]
    part_by_number = await _api_parts_by_number(filemaker, part_numbers)

    spare_numbers = list(dict.fromkeys(
        _text(_fields(record).get("替代編號"))
        or _text(part_by_number.get(_text(_fields(record).get("零件編號")), {}).get("替代編號"))
        or _text(part_by_number.get(_text(_fields(record).get("零件編號")), {}).get("spareparts"))
        for record in bom_records
    ))
    spare_numbers = [item for item in spare_numbers if item and item not in part_by_number][:500]
    if spare_numbers:
        part_by_number.update(await _api_parts_by_number(filemaker, spare_numbers))

    lines: list[CustomerBomLine] = []
    for record in bom_records:
        bom_fields = _fields(record)
        part_number = _text(bom_fields.get("零件編號"))
        part_fields = part_by_number.get(part_number, {})
        spare_number = (
            _text(bom_fields.get("替代編號"))
            or _text(part_fields.get("替代編號"))
            or _text(part_fields.get("spareparts"))
        )
        lines.append(CustomerBomLine(
            lineRef=str(record.get("recordId") or ""),
            partNumber=part_number,
            clientPartNumber=_text(part_fields.get("客戶編號")),
            partName=_english_text(part_fields.get(PART_NAME_FIELD) or part_fields.get("English Name")),
            bomQuantity=bom_fields.get("需求數量"),
            requiredQuantity=bom_fields.get("倉庫需求"),
            stock=(
                part_fields.get(PART_STOCK_FIELD)
                if part_fields.get(PART_STOCK_FIELD) not in (None, "")
                else part_fields.get("current_stock")
            ),
            status=_english_status(part_fields.get("status") or part_fields.get("狀態")),
            sparePartNumber=spare_number,
            spareStock=(
                part_by_number.get(spare_number, {}).get(PART_STOCK_FIELD)
                if part_by_number.get(spare_number, {}).get(PART_STOCK_FIELD) not in (None, "")
                else part_by_number.get(spare_number, {}).get("current_stock")
            ),
        ))
    return lines


async def _api_parts_by_number(
    filemaker: FileMakerClient,
    part_numbers: list[str],
) -> dict[str, dict[str, Any]]:
    unique_numbers = list(dict.fromkeys(item for item in part_numbers if item))[:500]
    if not unique_numbers:
        return {}
    result = await filemaker.find_records(
        PRODUCT_PART_LAYOUT,
        query=[{"part_number": f"=={number}"} for number in unique_numbers],
        limit=500,
    )
    mapped: dict[str, dict[str, Any]] = {}
    for record in result["data"]:
        part_fields = _fields(record)
        part_number = _text(part_fields.get("part_number"))
        if part_number and part_number not in mapped:
            mapped[part_number] = part_fields
    return mapped


async def _related_products(
    filemaker: FileMakerClient,
    part_number: str,
    client_id: str,
) -> list[CustomerRelatedProduct]:
    if not part_number:
        return []
    bom_result = await filemaker.find_records(
        PRODUCT_BOM_LAYOUT,
        query={"零件編號": f"=={part_number}"},
        limit=500,
    )
    product_skus = list(dict.fromkeys(
        _text(_fields(record).get("ID_產品編號"))
        for record in bom_result["data"]
        if _text(_fields(record).get("ID_產品編號"))
    ))
    if not product_skus:
        return []
    product_result = await filemaker.find_records(
        PRODUCT_LAYOUT,
        query=[
            {"product_sku": f"=={sku}", "id_client": f"=={client_id}"}
            for sku in product_skus
        ],
        limit=min(500, len(product_skus) * 2),
        sort=[{"fieldName": "product_sku", "sortOrder": "ascend"}],
    )
    related: list[CustomerRelatedProduct] = []
    seen: set[str] = set()
    for record in product_result["data"]:
        fields = _fields(record)
        sku = _text(fields.get("product_sku"))
        if not sku or sku not in product_skus or sku in seen:
            continue
        seen.add(sku)
        related.append(CustomerRelatedProduct(
            productRef=str(record.get("recordId") or ""),
            productSku=sku,
            productName=_english_text(fields.get("product_name")),
        ))
    return related


async def _order_details(
    filemaker: FileMakerClient,
    order_records: list[dict[str, Any]],
    web_client_id: str,
) -> dict[tuple[str, str], dict[str, Any]]:
    identities = list(dict.fromkeys(
        _order_identity(_fields(record))
        for record in order_records
        if any(_order_identity(_fields(record)))
    ))
    if not identities:
        return {}

    query: list[dict[str, str]] = []
    for internal_number, order_number in identities:
        criteria: dict[str, str] = {
            ORDER_SCOPE_FIELD: f"=={web_client_id.strip()}",
        }
        if internal_number:
            criteria[ORDER_INTERNAL_ID_FIELD] = f"=={internal_number}"
        if order_number:
            criteria["出貨單 PI"] = f"=={order_number}"
        if criteria:
            query.append(criteria)
    if not query:
        return {}

    result = await filemaker.find_records(
        LEGACY_ORDER_DETAIL_LAYOUT,
        query=query,
        limit=min(MAX_CATALOG_PAGE_SIZE * 2, max(len(query) * 2, len(query))),
    )
    details: dict[tuple[str, str], dict[str, Any]] = {}
    allowed_identities = set(identities)
    for record in result["data"]:
        fields = _fields(record)
        identity = _order_identity(fields)
        if identity in allowed_identities and identity not in details:
            details[identity] = fields
    return details


def _order_identity(fields: dict[str, Any]) -> tuple[str, str]:
    return (
        _text(fields.get(ORDER_INTERNAL_ID_FIELD)),
        _text(fields.get("出貨單 PI")),
    )


def _order(
    record: dict[str, Any],
    detail_fields: dict[str, Any],
    *,
    can_view_price: bool = True,
) -> CustomerCatalogOrder:
    fields = _fields(record)
    visible_fields = {**fields, **detail_fields}
    return CustomerCatalogOrder(
        orderRef=str(record.get("recordId") or ""),
        clientName=_text(visible_fields.get("出貨單_客戶::客戶名稱")),
        orderNumber=_public_order_number(fields, detail_fields),
        orderAmount=visible_fields.get(ORDER_AMOUNT_FIELD) if can_view_price else None,
        shippingCompany=_text(visible_fields.get("shipping_company")),
        trackingNumber=_text(visible_fields.get("tracking_number")),
        shippingCost=visible_fields.get("shipping_cost") if can_view_price else None,
        shippedDate=_text(visible_fields.get("出貨日期")),
        shippingStatus=_shipping_status(visible_fields),
        remarks=(
            _text(visible_fields.get("order_remarks_for_client_only"))
            or _text(visible_fields.get("shipping_notes"))
        ),
    )


def _public_order_number(
    fields: dict[str, Any],
    detail_fields: dict[str, Any],
) -> str:
    """Expose only the customer PO; FileMaker PI/internal numbers stay private."""
    order_po = _text(fields.get("訂單 PO")) or _text(detail_fields.get("訂單 PO"))
    if re.match(r"^PI(?:[-#\s]|$)", order_po, flags=re.IGNORECASE):
        return ""
    return order_po


def _shipping_status(fields: dict[str, Any]) -> Literal["Shipped", "Not Shipped"]:
    return "Shipped" if _text(fields.get("出貨日期")) else "Not Shipped"


def _product_inventory_workbook(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:D{max(1, len(records) + 1)}"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 36
    sheet.column_dimensions["C"].width = 42
    sheet.column_dimensions["D"].width = 18
    sheet.append([
        "产品编号 / SKU",
        "中文名称 / Chinese Name",
        "英文名称 / English Name",
        "库存 / Inventory",
    ])
    # Use opaque ARGB colors. Six-digit RGB values can be interpreted as
    # transparent by some mobile Excel/WPS viewers, hiding the white titles.
    header_fill = PatternFill(fill_type="solid", fgColor="FF1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for record in records:
        fields = _fields(record)
        row_number = sheet.max_row + 1
        sheet.row_dimensions[row_number].height = 30
        text_values = (
            _text(fields.get("product_sku")),
            _text(fields.get("產品名稱_中文")),
            _english_text(fields.get("product_name")),
        )
        for column, value in enumerate(text_values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.data_type = "s"
            if column in (2, 3):
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        stock = _excel_number(fields.get(PRODUCT_STOCK_FIELD))
        stock_cell = sheet.cell(row=row_number, column=4, value=stock)
        if isinstance(stock, str):
            stock_cell.data_type = "s"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _part_inventory_workbook(records: list[dict[str, Any]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Parts"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(1, len(records) + 1)}"
    widths = {
        "A": 24,
        "B": 42,
        "C": 18,
        "D": 16,
        "E": 16,
        "F": 18,
        "G": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.append([
        "Part No.",
        "Part Name",
        "Status",
        "Inventory",
        "Safety Stock",
        "Turnover",
        "Created",
    ])
    header_fill = PatternFill(fill_type="solid", fgColor="FF1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFFFF", bold=True)
        cell.fill = header_fill

    for record in records:
        part = _part(record)
        row_number = sheet.max_row + 1
        text_values = (
            part.part_number,
            part.part_name,
            part.status,
        )
        for column, value in enumerate(text_values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.data_type = "s"
        inventory = _excel_number(part.stock)
        inventory_cell = sheet.cell(row=row_number, column=4, value=inventory)
        if isinstance(inventory, str):
            inventory_cell.data_type = "s"
        safety_stock = _excel_number(part.safety_stock)
        safety_cell = sheet.cell(row=row_number, column=5, value=safety_stock)
        if isinstance(safety_stock, str):
            safety_cell.data_type = "s"
        for column, value in ((6, part.turnover), (7, part.created)):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.data_type = "s"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_number(value: Any) -> int | float | str | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    try:
        numeric = float(text.replace(",", ""))
    except ValueError:
        return text
    return int(numeric) if numeric.is_integer() else numeric


def _product(
    record: dict[str, Any],
    *,
    inventory_only: bool = False,
) -> CustomerCatalogProduct:
    fields = _fields(record)
    return CustomerCatalogProduct(
        productRef=str(record.get("recordId") or ""),
        productSku=_text(fields.get("product_sku")),
        productName=_english_text(fields.get("product_name")),
        modelName="" if inventory_only else _english_text(fields.get("車款")),
        scale="" if inventory_only else _english_text(fields.get("車子比例")),
        category="" if inventory_only else _english_text(fields.get("類別")),
        stock=fields.get(PRODUCT_STOCK_FIELD),
        bomCount=None if inventory_only else fields.get("BOM計數"),
        hasImage=False if inventory_only else bool(_text(fields.get("檔案 1 | 容器"))),
    )


def _product_detail(
    record: dict[str, Any],
    *,
    can_view_price: bool = True,
    price: Any = None,
) -> CustomerProductDetailItem:
    fields = _fields(record)
    resolved_price = price if price is not None else fields.get("產品售價::Price")
    financial_payload = (
        {
            "price": resolved_price,
            "stockValue": fields.get("Stock_USD"),
            "prepaidStock": fields.get("PrePaid_stock_USD"),
        }
        if can_view_price
        else {}
    )
    return CustomerProductDetailItem(
        **_product(record).model_dump(by_alias=True),
        soldTotal=fields.get("產品庫存::出庫數量總合"),
        productionCalculation=fields.get("下單數量"),
        **financial_payload,
    )


def _product_image(record: dict[str, Any]) -> CustomerProductImage:
    fields = _fields(record)
    return CustomerProductImage(
        assetRef=str(record.get("recordId") or ""),
        filename=_english_text(fields.get("original_filename")),
        title=_english_text(fields.get("title")),
        sortOrder=fields.get("sort_order") or 0,
        isPrimary=fields.get("is_primary") or False,
    )


def _part(
    record: dict[str, Any],
    *,
    inventory_only: bool = False,
) -> CustomerCatalogPart:
    fields = _fields(record)
    safety_stock = (
        fields.get(PART_SAFETY_STOCK_FIELD)
        if fields.get(PART_SAFETY_STOCK_FIELD) not in (None, "")
        else fields.get("Safty Stock QTY")
    )
    turnover = _english_text(fields.get("Turnover Time"))
    return CustomerCatalogPart(
        partRef=str(record.get("recordId") or ""),
        partNumber=_text(fields.get("part_number")),
        partName=_english_text(fields.get(PART_NAME_FIELD) or fields.get("English Name")),
        stock=(
            fields.get(PART_STOCK_FIELD)
            if fields.get(PART_STOCK_FIELD) not in (None, "")
            else fields.get("current_stock")
        ),
        safetyStock=None if inventory_only else 0 if safety_stock in (None, "") else safety_stock,
        turnover="" if inventory_only else turnover or "0 Days",
        created="" if inventory_only else _text(fields.get("Date Created")),
        status="" if inventory_only else _english_status(fields.get("status")),
        hasImage=(
            False
            if inventory_only
            else bool(_text(fields.get("影像 | 容器")) or _text(fields.get("圖面 | 容器")))
        ),
    )


def _fields(record: dict[str, Any]) -> dict[str, Any]:
    fields = record.get("fieldData")
    return fields if isinstance(fields, dict) else record


def _english_status(value: Any) -> str:
    text = _text(value)
    normalized = text.casefold().replace(" ", "")
    translations = {
        "待确认": "Pending confirmation",
        "待確認": "Pending confirmation",
        "已确认": "Confirmed",
        "已確認": "Confirmed",
        "进行中": "In progress",
        "進行中": "In progress",
        "完成": "Completed",
        "暂停": "On hold",
        "暫停": "On hold",
        "停用": "Inactive",
        "正常": "Active",
    }
    if normalized in translations:
        return translations[normalized]
    return "" if re.search(r"[\u3400-\u9fff]", text) else text


def _english_text(value: Any) -> str:
    text = _text(value)
    translations = {
        "零件包": "Parts kit",
        "成品": "Finished product",
        "整车": "Complete vehicle",
        "整車": "Complete vehicle",
        "配件": "Accessories",
    }
    if text in translations:
        return translations[text]
    if re.search(r"[\u3400-\u9fff]", text):
        return ""
    cleaned = re.sub(r"[\u3040-\u30ff]+", "", text)
    return re.sub(r"\s{2,}", " ", cleaned).strip()


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _not_found(entity: str) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail={"message": f"{entity} not found."},
    )


def _catalog_unavailable() -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_502_BAD_GATEWAY,
        detail={"message": "The catalog service is temporarily unavailable. Please try again later."},
    )
