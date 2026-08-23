"""Verify deployed Stock Check product and part inventory Excel exports."""

from __future__ import annotations

from io import BytesIO
import json

import httpx
from openpyxl import load_workbook

from app.core.config import Settings
from app.services.customer_chat_auth import issue_customer_token, load_customer_accounts


def main() -> None:
    settings = Settings()
    accounts = load_customer_accounts(settings)
    account = next((item for item in accounts.values() if item.is_admin), next(iter(accounts.values())))
    token, _ = issue_customer_token(account, settings)
    headers = {"Authorization": f"Bearer {token}"}

    results = []
    exports = (
        ("products", "Products", ("SKU", "Inventory")),
        (
            "parts",
            "Parts",
            ("Part No.", "Part Name", "Status", "Inventory", "Safety Stock", "Turnover", "Created"),
        ),
    )
    with httpx.Client(base_url="http://frontend", headers=headers, timeout=120) as client:
        for catalog, sheet_name, expected_columns in exports:
            response = client.get(f"/api/customer-chat/catalog/{catalog}/export.xlsx")
            response.raise_for_status()
            assert response.headers["content-type"].startswith(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            assert response.headers["content-disposition"].endswith('.xlsx"')
            exported_count = int(response.headers["x-export-row-count"])
            workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            assert rows[0] == expected_columns
            assert len(rows) == exported_count + 1
            assert all(len(row) == len(expected_columns) for row in rows)
            assert all(row[0] for row in rows[1:])
            results.append({
                "catalog": catalog,
                "status": response.status_code,
                "contentType": response.headers["content-type"],
                "file": response.headers["content-disposition"],
                "exportedRows": exported_count,
                "worksheetRows": len(rows),
                "columns": list(rows[0]),
                "bytes": len(response.content),
            })

    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
